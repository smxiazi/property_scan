#!/usr/bin/env python
# coding:utf-8
from urllib.parse import urlparse
import dns.resolver
import os,requests,json,socket,time,csv,re,urllib3
from bs4 import BeautifulSoup
import subprocess
import ipaddress
import shutil
import dns.resolver
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
import threading,queue,base64
from http.server import HTTPServer, BaseHTTPRequestHandler

user = "admin" #登录账号
passwd = "admin123" #登录密码
web_port = 8000 #web界面端口

root_domain_list = [] #存放白名单根域
all_domain_list = {} #存放符合根域的子域名
all_domain_url_list = {} #存放符合根域的子域名URL
all_ip_list = [] #存放ip
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning) #禁用requests的https证书告警
all_time_list = [0,0,0] #程序运行时间记录 子域名资产、子域名web资产、互联网资产
save_file = "output/"+str(int(time.time()))+".xlsx" #文件保存路径
log_data = [] #存储日志


#接收域名处理 传参：【url、根域名】     #子域名web资产
def input_domain(url,root_domain):
    # 根域名、子域名、IP、物理地址、链接、响应码、title、端口、协议、中间件
    if url.startswith('http://') or url.startswith('https://'):
        try:
            parsed_url = urlparse(url)
            protocol = parsed_url.scheme #协议 http、https
            domain = parsed_url.netloc.split(":")[0]  # 获取子域名
            port = parsed_url.port #端口，如果没有带端口默认回味None
            if port is None:# 如果端口味None，则根据协议来填写端口
                # 如果不包含端口，根据协议返回默认端口
                if parsed_url.scheme == 'http':
                    port = 80
                elif parsed_url.scheme == 'https':
                    port = 443

            if is_port_open(domain,port) == False:
                print("网站关闭，端口无法访问")
                return []

            #获取标题和中间件信息
            new_url = protocol+"://"+domain+":"+str(port)
            title_data =  get_website_title(new_url)
            if len(title_data) != 0:
                title = title_data[0]
                head_Server = title_data[1]
                status_code = title_data[2]
            else:
                title = ""
                head_Server = ""
                status_code = ""

            #获取域名的ip地址
            try:
                ip_address = socket.gethostbyname(domain) #获取域名的ip地址
                ip_address_location = get_ip_address_location(ip_address)[0]
            except socket.gaierror:
                print("获取ip地址失败")
                ip_address = ""
                ip_address_location = ""
            return [root_domain, domain,ip_address,ip_address_location,new_url,status_code,title,port,protocol,head_Server]
        except Exception as e:
            print("处理域名时错误：",e)
            return []
    else:
        print("url中没有带协议")
        return []

#接收ip处理 传参：【ip】    #互联网服务器资产
def input_ip(ip):
    #IP、物理地址、链接、title、端口、协议、中间件
    list_port = []
    ip_address_location = get_ip_address_location(ip)
    masscan_data = masscan_port_scan(ip)
    if len(masscan_data) != 0:
        if len(masscan_data[1]) != 0:
            port_servic = nmap_servic_scan(ip,masscan_data[1])
            for i in port_servic:
                if "ssl/http" in i[1]:
                    url = "https://"+ip+":"+i[0]
                    title_data = get_website_title(url)
                    if len(title_data)!= 0 :
                        list_port.append([ip, ip_address_location[0], url,title_data[0],i[0], i[1],title_data[1]])
                    else:
                        list_port.append([ip, ip_address_location[0], "", "", i[0], i[1], ""])
                elif "http" in i[1]:
                    url = "http://" + ip + ":" + i[0]
                    title_data = get_website_title(url)
                    if len(title_data)!= 0 :
                        list_port.append([ip, ip_address_location[0], url,title_data[0],i[0], i[1],title_data[1]])
                    else:
                        list_port.append([ip, ip_address_location[0], "", "", i[0], i[1], ""])
                else:
                    list_port.append([ip, ip_address_location[0], "", "", i[0], i[1], ""])
                    #list_port.append([ip, ip_address_location[0], "链接", "title", i[0], i[1], "中间件"])
            for i in list_port:
                print(i)
            return list_port
        else:
            return []
    else:
        return []

#dns请求，A记录和CNAME记录 传参：【域名、根域名】    #子域名资产
def get_dns_record(domain,root_domain):
    #根域名、子域名、记录类型、IP/别名
    try:
        # 使用 resolve() 方法，但检查返回的 rdata 是否为 CNAME
        answers = dns.resolver.resolve(domain, 'CNAME')  # 注意：这实际上会尝试解析 CNAME 链
        cname_found = False
        CNAME_data = ""
        for rdata in answers:
            if rdata.rdtype == dns.rdatatype.CNAME:
                CNAME_data = rdata.target.to_text() # CNAME 解析记录
                cname_found = True
        if cname_found:
            return [root_domain,domain,'CNAME',CNAME_data]
    except dns.resolver.NoAnswer:
        # 尝试获取 A 记录作为备选
        try:
            A_ip = ""
            answers = dns.resolver.resolve(domain, 'A')
            for rdata in answers:
                A_ip = rdata.address #A记录解析，只获取最终的那个ip
            return [root_domain,domain,'A',A_ip]
        except dns.resolver.NoAnswer:
            print(f"没有找到解析记录：{domain}")
            return []
    except dns.resolver.NXDOMAIN as e:
        print(f"域名不存在： {domain} ",e)
        return []

#masscan全端口扫描
def masscan_port_scan(ip):
    port_list = []
    command = f"masscan -p 1-65535 --rate 1000 {ip}"
    result = subprocess.run(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    if result.returncode == 0:
        data = result.stdout.split("\n")
        for i in data:
            if "Discovered open port" in i:
                i = i.split(' ')
                try:
                    #p_i = i2[5] + ":" + i2[3].split('/')[0]
                    port_list.append(i[3].split('/')[0])#端口
                except:
                    pass
        if len(port_list) < 50:
            print([ip,port_list])
            return [ip,port_list]
        else:
            print("开放端口大于50个，丢弃")
            return []
    else:
        print("Error:", result.stderr)
        return []

#nmap端口服务指纹识别
def nmap_servic_scan(ip,port_list):
    port_servi_list = []
    port_list_str = [str(port) for port in port_list]  # 将每个整数转换成字符串
    port = ','.join(port_list_str)
    t1 = time.time()
    command = f"nmap -p {port} {ip} -sV --version-intensity 1 -T4 --host-timeout 600s --data-length 8"
    result = subprocess.run(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    if result.returncode == 0:
        data = result.stdout.split("\n")
        for i in data:
            if " open  " in i:
                i = i.split(" ")
                new_i = [item for item in i if item != ""]# 使用列表推导式删除空字符串
                port = new_i[0].split("/")[0]#端口
                servic = new_i[2]#服务
                port_servi_list.append([port, servic])
        print(port_servi_list)
        return port_servi_list
    else:
        print("Error:", result.stderr)
    print("nmap识别服务花费：" + str(int(time.time() - t1)) + "秒")
    return []

#获取ip归属地
def get_ip_address_location(ip):
    count = 0
    while True:
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:109.0) Gecko/20100101 Firefox/114.0'
            }
            get_address_url = 'https://api3.toolnb.com/tools/ipgetareainfo.json?ip=' + ip
            web_data = requests.get(get_address_url, headers=headers).text
            if "出错" in web_data:
                print(web_data)
                return [""]
            ip_address_location = json.loads(web_data)['data']['area']  # 获取ip地址的归宿地
            return [ip_address_location]
        except Exception as e:
            print("获取ip地址失败,尝试再次获取", e)
            if count == 3:
                return [""]
            count += 1

#更新网站标题、中间件信息
def get_website_title(url):
    try:
        # 发送 HTTP GET 请求
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:109.0) Gecko/20100101 Firefox/114.0'
        }
        response = requests.get(url,headers=headers,timeout=5,verify=False)

        server_header = response.headers.get('Server')
        if server_header: #中间件信息
            server_header = server_header
        else:
            server_header = ""

        # 尝试从HTTP响应头中获取编码（如果有的话）
        encoding = response.encoding if 'charset' in response.headers.get('content-type','').lower() else None

        # 如果响应头中没有编码信息，则尝试从HTML文档的<meta charset>标签中获取
        if not encoding:
            soup = BeautifulSoup(response.content, 'html.parser')
            meta_charset = soup.find('meta', attrs={'charset': True})
            if meta_charset and 'content' in meta_charset.attrs:
                encoding = meta_charset['content']

        # 如果还没有编码信息，则假设为UTF-8（这是现代网页最常见的编码）
        if not encoding:
            encoding = 'utf-8'

        # 使用找到的编码解码响应内容
        response.encoding = encoding

        # 使用BeautifulSoup解析HTML并获取标题
        soup = BeautifulSoup(response.text, 'html.parser')
        title = soup.title.string if soup.title else ''
        return [title,server_header,response.status_code]
    except requests.exceptions.RequestException as e:
        # 处理请求异常
        print(f"获取网站标题错误: {e}")
        return []

#判断网站端口是否开放
def is_port_open(host, port):
    try:
        # 创建一个socket对象
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        # 设置超时时间，防止无限等待
        sock.settimeout(5)
        # 尝试连接到目标主机的指定端口
        result = sock.connect_ex((host, port))
        # 如果连接成功，connect_ex()返回0
        if result == 0:
            #print(f"{host}:{port} 是开放的")
            return True
        else:
            #print(f"{host}:{port} 是关闭的")
            return False
    except Exception as e:
        #print(f"发生错误: {e}")
        return False
    finally:
        # 关闭socket连接
        sock.close()

#高危端口
def high_port():
    high_port_list = [21,22,23,25,53,109,110,135,137,138,139,134,161,445,512,513,873,1433,1434,1521,3306,3389,5432,6379,7001,7002,9080,9090,9200,9300]
    return high_port_list

#辅助方法-寻找字符的下标
def find_nth_occurrence(string, char, n):
    index = -1
    for _ in range(n):
        index = string.find(char, index + 1)
        if index == -1:
            return -1  # 如果找不到第 n 个出现的字符，返回 -1
    return index

#记录器-获取列表最新数据量，用于判断数据是否有更新
def get_count():
    all_domain_list_count = {}
    all_domain_url_list_count = {}
    all_ip_list_count = len(all_ip_list)
    for root_domain in root_domain_list:
        data_tmp = all_domain_list.get(root_domain)
        if data_tmp != None:
            all_domain_list_count[root_domain] = len(data_tmp)
        else:
            all_domain_list_count[root_domain] = 0

        data_tmp = all_domain_url_list.get(root_domain)
        if data_tmp != None:
            all_domain_url_list_count[root_domain] = len(data_tmp)
        else:
            all_domain_url_list_count[root_domain] = 0

    # print("\n\n最新数据：")
    # for key,value in all_domain_list_count.items():
    #     print("子域名数量：",key,value)
    #
    # for key,value in all_domain_url_list_count.items():
    #     print("子域url名数量：",key,value)
    #
    # print("ip数量：",all_ip_list_count)
    return [all_domain_list_count,all_domain_url_list_count,all_ip_list_count]

#通过输入秒数，分析出 小时、分钟、秒
def seconds_to_hms(seconds):
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return str(int(h)), str(int(m)), str(int(s))

#扫描结果保存到本地xlsx文件
def xlsx_save(list,index):
    # 加载现有的工作簿
    workbook = load_workbook(save_file)
    print(list)
    if index == 1: #子域名资产
        # 选择活动工作表，或者通过名称选择工作表
        sheet = workbook['子域名资产']  # 使用名为'Sheet1'的工作表
        sheet.append(list)
        # 保存工作簿
        workbook.save(save_file)  # 这会覆盖原始文件
    elif index == 2:#子域名web资产
        # 选择活动工作表，或者通过名称选择工作表
        sheet = workbook['子域名web资产']  # 使用名为'Sheet1'的工作表
        sheet.append(list)
        # 保存工作簿
        workbook.save(save_file)  # 这会覆盖原始文件
    elif index == 3:#ip资产
        # 选择活动工作表，或者通过名称选择工作表
        sheet = workbook['互联网服务器资产']  # 使用名为'Sheet1'的工作表
        for i in list:
            if int(i[4]) in high_port():
                i.append("高危端口")
                sheet.append(i)
                # 获取最后一行的索引（注意，索引从1开始）
                last_row = sheet.max_row
                # 遍历最后一行的每个单元格，设置背景色为红色
                for col in sheet.iter_cols(min_row=last_row, max_row=last_row, min_col=1, max_col=sheet.max_column):
                    for cell in col:
                        # 创建一个填充样式，这里使用红色背景
                        fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                        # 应用样式到单元格
                        cell.fill = fill
            else:
                sheet.append(i)
        workbook.save(save_file)  # 这会覆盖原始文件

#过滤器-判断是否扫描过
def domai_or_ip_filtering(data_list):
    for i in data_list:
        i = i.strip("\n")
        i = i.strip("\r")
        if i.startswith("https://") or i.startswith("http://"):
            #print("\n是web协议",i)
            x_index = find_nth_occurrence(i,"/",3) #第三个/字符的下标
            # 去掉uri的信息
            if x_index != -1:
                domain = i[0:x_index]
            else:
                domain = i
            m_index_1 = find_nth_occurrence(domain,":",1) #第1个:字符的下标
            m_index_2 = find_nth_occurrence(domain,":",2) #第2个:字符的下标
            # 去掉后面带端口的信息
            if m_index_2 != -1:
                domain_2 = domain[m_index_1+3:m_index_2]
                if domain[m_index_2+1:] == "80" or domain[m_index_2+1:] == "443":#如果:号后面带的端口是80或443则删除掉
                    domain = i[0:m_index_2]
            else:
                domain_2 = domain[m_index_1+3:]

            #print(domain, domain_2)

            try:
                ip_obj = ipaddress.ip_address(domain_2)#判断是否是ip
                # 判断 IP 是否是私有地址
                if ip_obj.is_private:
                    print("内网ip：", domain_2)
                    continue
                else:
                    #print("不是内网ip：", domain_2)
                    if domain_2 not in all_ip_list:
                        all_ip_list.append(domain_2)
                    else:
                        print("该ip已存在：",domain_2)


            except:
                for root_domain in root_domain_list:
                    if domain_2.endswith("."+root_domain):
                        #print("域名在范围内：",domain)

                        # 存放子域名
                        try:
                            domain_tmep_list = all_domain_list[root_domain]
                            if domain_2 not in domain_tmep_list:
                                domain_tmep_list.append(domain_2)
                                all_domain_list[root_domain] = domain_tmep_list
                            else:
                                print("该子域名已存在：", domain_2)
                        except:
                            all_domain_list[root_domain] = [domain_2]


                        try:
                            domain_tmep_list = all_domain_url_list[root_domain]
                            if domain not in domain_tmep_list:
                                domain_tmep_list.append(domain)
                                all_domain_url_list[root_domain] = domain_tmep_list
                            else:
                                print("该URL已存在：", domain)
                        except:
                            all_domain_url_list[root_domain] = [domain]
                        break

        else:
            #print("\n他的不开头是web协议！！！",i)
            domain = i.split("/")[0]
            domain = domain.split(":")[0]
            pattern = re.compile(r'^[0-9a-zA-Z\.]+$')
            if bool(pattern.match(domain)) == False:#判断字符串是否符合域名的要求
                continue
            try:
                ip_obj = ipaddress.ip_address(domain)#判断是否是ip
                # 判断 IP 是否是私有地址
                if ip_obj.is_private:
                    print("内网ip：", domain)
                    continue
                else:
                    #print("不是内网ip：", domain)
                    if domain not in all_ip_list:
                        all_ip_list.append(domain)
                    else:
                        print("该ip已存在：",domain)
            except:
                for root_domain in root_domain_list:
                    if domain.endswith("."+root_domain):
                        #print("域名在范围内：",domain)


                        #存放子域名
                        try:
                            domain_tmep_list = all_domain_list[root_domain]
                            if domain not in domain_tmep_list:
                                domain_tmep_list.append(domain)
                                all_domain_list[root_domain] = domain_tmep_list
                            else:
                                print("该子域名已存在：", domain)
                        except:
                            all_domain_list[root_domain] = [domain]

                        #存放域名的url地址
                        for i in ["http://","https://"]:
                            try:
                                domain_tmep_list = all_domain_url_list[root_domain]
                                if i + domain not in domain_tmep_list:
                                    domain_tmep_list.append(i+domain)
                                    all_domain_url_list[root_domain] = domain_tmep_list
                                else:
                                    print("该URL已存在：", i+domain)
                            except:
                                all_domain_url_list[root_domain] = [i+domain]

                        break

#主程序逻辑
def go_run():
    old_count =[{},{},0]

    domain_dns_ip = [] #用来存放子域名解析出来的ip

    try:
        directory = Path('output')
        directory.mkdir(parents=True, exist_ok=True)
        # 尝试复制文件
        shutil.copy("web/模板.xlsx",save_file)
        print(f"文件成功复制到 {save_file}")
    except FileNotFoundError:
        print(f"源文件 模板.xlsx 不存在")
    except Exception as e:
        print(f"复制文件时发生错误: {e}")

    while True:
        print("=======================================================")
        print_data = [] #用来保存打印的内容，为什么这样写，因为要控制打印的位置
        new_count = get_count()

        is_run = True #用来处理优先级，如果有任务，先跑 子域名资产 - 》 url的web资产 -》 ip资产
        #子域名解析
        is_run_domain = False
        for key, value in new_count[0].items():
            if old_count[0].get(key) != None:
                if old_count[0].get(key) < value and is_run_domain == False:
                    start_time = time.time() #开始时间
                    print(all_domain_list.get(key)[old_count[0].get(key)],key)
                    data_domain_dns = get_dns_record(all_domain_list.get(key)[old_count[0].get(key)],key)
                    #保存到xlsx中
                    if len(data_domain_dns) != 0 :
                        xlsx_save(data_domain_dns,1)
                    #print(data_domain_dns)

                    #添加dns解析A记录的ip地址，用于后续的 服务器资产扫描
                    if len(data_domain_dns) != 0:
                        if data_domain_dns[2] == "A":
                            domain_dns_ip.append(data_domain_dns[3])

                    old_count[0][key] = old_count[0][key]+1
                    is_run_domain = True
                    is_run = False
                    end_time = time.time() #结束时候
                    all_time_list[0] = all_time_list[0] + (end_time-start_time)#使用时间累计
                if old_count[0].get(key) == value:
                    print_data.append([key,"子域名数量：",value,"目前处理的进度是：",old_count[0].get(key),"完成"])
                else:
                    print_data.append([key, "子域名数量：", value, "目前处理的进度是：", old_count[0].get(key)])

            else:
                #初始化
                old_count[0][key] = 0
                print_data.append([key,"子域名数量：",value,"目前处理的进度是：",old_count[0].get(key)])
                is_run = False

        #用来处理子域名解析出来的ip
        if len(domain_dns_ip) !=0 and is_run == True:
            domain_dns_ip = list(set(domain_dns_ip)) #去重ip
            domai_or_ip_filtering(domain_dns_ip)
            domain_dns_ip = []

        #子域名url
        is_run_domain_url = False
        for key, value in new_count[1].items():
            if old_count[1].get(key) != None:
                if old_count[1].get(key) < value and is_run_domain_url == False and is_run == True:
                    start_time = time.time()  # 开始时间
                    print(all_domain_url_list.get(key)[old_count[1].get(key)],key)
                    data_domain = input_domain(all_domain_url_list.get(key)[old_count[1].get(key)],key)
                    #print(data_domain)
                    # 保存到xlsx中
                    if len(data_domain) != 0 :
                        xlsx_save(data_domain,2)

                    old_count[1][key] = old_count[1][key]+1
                    is_run_domain_url = True
                    is_run = False
                    end_time = time.time()  # 结束时候
                    all_time_list[1] = all_time_list[1] + (end_time - start_time) #使用时间累计
                if old_count[1].get(key) == value:
                    print_data.append([key,"子域名url数量：", value,"目前处理的进度是：",old_count[1].get(key),"完成"])
                else:
                    print_data.append([key,"子域名url数量：", value,"目前处理的进度是：",old_count[1].get(key)])

            else:
                # 初始化
                old_count[1][key] = 0
                print_data.append([key,"子域名url数量：", value,"目前处理的进度是：",old_count[1].get(key)])

        #ip
        if old_count[2] < new_count[2] and is_run == True:
            start_time = time.time()  # 开始时间
            print(all_ip_list[old_count[2]])
            data_ip = input_ip(all_ip_list[old_count[2]])
            #print(data_ip)
            # 保存到xlsx中
            if len(data_ip) != 0:
                xlsx_save(data_ip, 3)
            old_count[2] = old_count[2]+1
            end_time = time.time()  # 结束时候
            all_time_list[2] = all_time_list[2] + (end_time - start_time)#使用时间累计
            print_data.append(["ip数量：",new_count[2],"目前处理的进度是：",old_count[2]])
        else:
            if old_count[2] == new_count[2]:
                print_data.append(["ip数量：",new_count[2],"目前处理的进度是：",old_count[2],"完成"])
            else:
                print_data.append(["ip数量：", new_count[2], "目前处理的进度是：", old_count[2]])

        time_class = ["子域名解析", "子域名web资产", "ip"]
        print_data.append(["----------------------------------------------"])
        for i in range(0, 3):
            hours, minutes, seconds = seconds_to_hms(all_time_list[i])
            log_time = f"{time_class[i]} 处理累计耗时：{hours}小时{minutes}分钟{seconds}秒"
            print_data.append([log_time])
            #print(log_time)
        print("----------------------------------------------")

        if old_count == new_count:
            print_data.append(["----------------------------------------------\n任务已全部完成，请查收！！！！"])
            for i in print_data:
                print(" ".join([str(data_str) for data_str in i]))
            time.sleep(10)
        else:
            print("----------------------------------------------")
            print_data.append([old_count, "\n", new_count]) #当前任务进展
            for i in print_data:
                print(" ".join([str(data_str) for data_str in i]))

            #time.sleep(2)
        global log_data
        log_data = print_data

#web页面
class SimpleHTTPRequestHandler(BaseHTTPRequestHandler):

    def do_GET(self):
        """处理GET请求"""
        # 检查是否提供了认证头
        auth = self.headers.get('Authorization')
        if not auth:
            self.send_response(401)
            self.send_header('WWW-Authenticate', 'Basic realm="Secure Area"')
            self.end_headers()
            self.wfile.write(b'No auth header received')
            return

        # 验证认证信息
        auth_decoded = base64.b64decode(auth.split()[1]).decode('utf-8')
        username, password = auth_decoded.split(':')

        if username == user and password == passwd:
            if self.path == '/':
                if os.path.exists("web/index.html"):
                    with open("web/index.html", 'r') as file:
                        index_data = ""
                        for index_read in file.readlines():
                            if "#####" == index_read.strip("\n"):
                                for i in root_domain_list:
                                    index_data += f"<tr><td>{i}</td></tr>"

                            else:
                                index_data += index_read

                        self.send_response(200)
                        self.send_header('Content-type', 'text/html')
                        self.end_headers()
                        self.wfile.write(index_data.encode("utf-8"))

            elif self.path == '/log.html':
                self.path = "/web" + self.path
                file_path = os.path.join(os.getcwd(), self.path[1:])  # 假设静态文件在当前工作目录下
                html_data = ""
                if os.path.exists(file_path):
                    with open(file_path, 'r') as file:
                        for i in file.readlines():
                            if "####" == i.strip("\n"):
                                for log in log_data:
                                    log = " ".join([str(log_data) for log_data in log])
                                    if "\n" in log:
                                        log = log.replace("\n","<p>")
                                    log = "<li>" + log + "</li>"
                                    html_data += log
                            else:
                                html_data += i


                        self.send_response(200)
                        self.send_header('Content-type', 'text/html')
                        self.end_headers()
                        self.wfile.write(html_data.encode("utf-8"))

            elif self.path == '/download':
                file_path = save_file + "_download"
                try:
                    # 尝试复制文件
                    shutil.copy(save_file,file_path)
                    print(f"文件成功复制到 {file_path}")
                except FileNotFoundError:
                    print(f"源文件 {save_file} 不存在")
                except Exception as e:
                    print(f"复制文件时发生错误: {e}")

                # 提供文件下载
                if os.path.exists(file_path):
                    self.send_response(200)
                    self.send_header('Content-type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    self.send_header('Content-Disposition', 'attachment; filename="output.xlsx"')
                    with open(file_path, 'rb') as file:
                        self.send_header('Content-Length', str(os.path.getsize(file_path)))
                        self.end_headers()
                        self.wfile.write(file.read())
                else:
                    self.send_error(404, 'File Not Found')
                try:
                    # 尝试删除文件
                    os.remove(file_path)
                    print(f"文件 {file_path} 已成功删除")
                except FileNotFoundError:
                    print(f"文件 {file_path} 不存在")
                except Exception as e:
                    print(f"删除文件时发生错误: {e}")

            elif  self.path.endswith('.js') or self.path.endswith('.css') or self.path.endswith('.svg') or self.path.endswith('.ttf') or self.path.endswith('.woff') or self.path.endswith('.woff2') or self.path.endswith('.eot'):
                # 处理JS、css文件
                self.path = "/web" + self.path
                file_path = os.path.join(os.getcwd(), self.path[1:])  # 假设静态文件在当前工作目录下
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as file:
                        self.send_response(200)
                        if self.path.endswith('.js'):
                            self.send_header('Content-type', 'application/javascript')
                        elif self.path.endswith('.css'):
                            self.send_header('Content-type', 'text/css')
                        else:
                            self.send_header('Content-type', 'application')
                        self.end_headers()
                        self.wfile.write(file.read())

            elif self.path.endswith('.html'):
                self.path = "/web" + self.path
                file_path = os.path.join(os.getcwd(), self.path[1:])  # 假设静态文件在当前工作目录下
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as file:
                        self.send_response(200)
                        self.send_header('Content-type', 'text/html')
                        self.end_headers()
                        self.wfile.write(file.read())

            else:
                # 其他路径，可以返回404或其他错误
                self.send_error(404, 'Not Found')
        else:
            self.send_response(401)
            self.send_header('WWW-Authenticate', 'Basic realm="Secure Area"')
            self.end_headers()
            self.wfile.write(b'Unauthorized')

    def do_POST(self):
        """处理POST请求"""
        # 检查是否提供了认证头
        auth = self.headers.get('Authorization')
        if not auth:
            self.send_response(401)
            self.send_header('WWW-Authenticate', 'Basic realm="Secure Area"')
            self.end_headers()
            self.wfile.write(b'No auth header received')
            return

            # 验证认证信息
        auth_decoded = base64.b64decode(auth.split()[1]).decode('utf-8')
        username, password = auth_decoded.split(':')

        if username == user and password == passwd:
            if self.path == '/api_root_domain':
                try:
                    content_length = int(self.headers['Content-Length'])
                    post_data = self.rfile.read(content_length)
                    data = post_data.decode('utf-8')
                    data = json.loads(data)
                    if data["domain"] not in root_domain_list:
                        root_domain_list.append(data["domain"])
                        rp_data = b'{"Success":"ok"}'
                    else:
                        rp_data = b'{"Success":"repetition!!"}'
                except:
                    rp_data = b'{"Success":"error!"}'
                self.send_response(200)
                self.send_header('Content-type', 'text/json')
                self.end_headers()
                self.wfile.write(rp_data)

            elif self.path == '/add_data':
                try:
                    content_length = int(self.headers['Content-Length'])
                    post_data = self.rfile.read(content_length)
                    data = post_data.decode('utf-8')
                    data = json.loads(data)
                    data_list = data["data"].split("\n")
                    domai_or_ip_filtering(data_list)
                    rp_data = b'{"Success":"ok"}'
                except:
                    rp_data = b'{"Success":"error!"}'
                self.send_response(200)
                self.send_header('Content-type', 'text/json')
                self.end_headers()
                self.wfile.write(rp_data)
            else:
                self.send_error(404, 'Not Found')
        else:
            self.send_response(401)
            self.send_header('WWW-Authenticate', 'Basic realm="Secure Area"')
            self.end_headers()
            self.wfile.write(b'Unauthorized')

#启动web界面
def web_run(server_class=HTTPServer, handler_class=SimpleHTTPRequestHandler):
    server_address = ('', web_port)
    httpd = server_class(server_address, handler_class)
    print('Starting httpd...')
    httpd.serve_forever()


def main():
    t1 = threading.Thread(target=go_run)
    t2 = threading.Thread(target=web_run)
    # 启动线程
    t1.start()
    t2.start()
    # 等待线程完成（主线程会继续执行）
    t1.join()
    t2.join()

main()
